import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        'Este script requiere openpyxl. Instálalo con: python -m pip install openpyxl'
    ) from exc

WORKBOOK_PATH = Path(__file__).resolve().parent / 'RIFA.xlsx'
OUTPUT_PATH = Path(__file__).resolve().parent / 'data_export.json'


def normalize_number(value):
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None
    digits = ''.join(ch for ch in value if ch.isdigit())
    if not digits:
        return None
    return int(digits)


def extract_numbers_from_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for row in rows:
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if text and text.replace(' ', '').isdigit():
                number = normalize_number(text)
                if number and 1 <= number <= 160:
                    result.append(number)
    return sorted(set(result))


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f'No se encontró el archivo: {WORKBOOK_PATH}')

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb.active
    extracted = extract_numbers_from_sheet(sheet)

    data = []
    for n in range(1, 161):
        data.append({
            'number': n,
            'name': '',
            'phone': '',
            'value': 0,
            'downPayment': 0,
            'status': 'available',
            'notes': '',
            'date': ''
        })

    for index, number in enumerate(extracted, start=1):
        if 1 <= number <= 160:
            data[number - 1]['status'] = 'reserved'
            data[number - 1]['name'] = f'Importado {index}'
            data[number - 1]['value'] = 20000
            data[number - 1]['downPayment'] = 10000
            data[number - 1]['date'] = '2026-08-17'

    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Se exportó la estructura de la rifa a {OUTPUT_PATH}')
    print(f'Número de entradas: {len(data)}')
    print(f'Números detectados: {len(extracted)}')


if __name__ == '__main__':
    main()
